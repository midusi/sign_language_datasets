from pathlib import Path
from os import path as osp, rename


class DatasetHandler(object):
    handler_class = {f'LSA64_raw': 'DH_LSA64',
                     'LSA64_cut': 'DH_LSA64',
                     'LSA64_pre': 'DH_LSA64_pre',
                     'ASLLVD_pre': 'DH_ASLLVD_pre'
                     }

    def __init__(self, version):
        self.version = version

    def get_my_path(self, root=None):
        # if datasets_path was not specified, use default
        if root is None:
            root = self.get_lib_root()
        return osp.join(root, self.version)

    def get_lib_root(self):
        return osp.join(Path.home(), '.sldatasets')

    @staticmethod
    def factory(version):
        try:
            e = DatasetHandler.handler_class[version]
            return globals()[e](version)
        except:
            raise ValueError(
                'version for lsa64("version"), must be "cut", "raw" or "pre" \
                version for asllvd must be None')

    def get_my_url(self):
        raise NotImplementedError

    def get_my_folder(self):
        raise NotImplementedError

    def get_my_file_ext(self):
        raise NotImplementedError

    def redux(self, files, **kwargs):
        l = list(filter(lambda f: osp.splitext(f)[1].endswith(
            f'{self.get_my_file_ext()}'), sorted(files)))

        def file_filter(self, f, clas, consultant, repetition):
            d_specs = self.video_info(f)
            try:
                clas=clas.lower()
                consultant=consultant.lower()            
            finally:
                return ((self.word_class(d_specs) == clas if clas else True) and
                        (self.consultant_of(d_specs) == consultant if consultant else True) and
                        (self.rept(d_specs) == repetition if repetition else True))
        l = list(filter(lambda f: file_filter(
            self, f, kwargs.get('word'), kwargs.get('consultant'), kwargs.get('repetition')), l))
        return l

    def specs_from(self, filename):
        d=self.video_info(filename)
        d['filename']= filename
        return d

    def video_info(self, filename):
        return NotImplementedError

    def word_count(self, videos_list):
        class_s = set([])
        consultant_s = set([])
        repetition_s = set([])
        for video in filter(lambda f: osp.splitext(f)[1].endswith(f'{self.get_my_file_ext()}'), videos_list):
            d = self.specs_from(video)
            class_s.add(self.word_class(d))
            consultant_s.add(self.consultant_of(d))
            repetition_s.add(self.rept(d))
        return [len(class_s), len(consultant_s), len(repetition_s)]

    def consultant_of(self, d):
        return d['consultant']

    def word_class(self, d):
        return d['class']

    def rept(self, d):
        return d['repetition']
    
    def get_pos_url(self):
        pass


class DH_LSA64(DatasetHandler):

    def get_my_url(self):
        if self.version == 'LSA64_raw':
            return 'https://drive.google.com/uc?id=1C7k_m2m4n5VzI4lljMoezc-uowDEgIUh'
        else:
            return 'https://drive.google.com/uc?id=18VuWBAxHaSBbO7wx57kQVre78FN7GYzQ'

    def get_pos_url(self):
        if self.version == 'LSA64_raw':
            raise NotImplementedError
        else:
            return 'https://drive.google.com/uc?id=1_byVg8q_GmfPvHps5irGRikEfeuYo5-r&'

    def get_my_folder(self):
        if self.version == 'LSA64_raw':
            return 'all'
        else:
            return 'all_cut'

    def get_my_file_ext(self):
        return 'mp4'

    def video_info(self,filename):
        splited = filename.split('_')
        info_dict={}
        info_dict['class']= int(splited[0])
        info_dict['consultant']= int(splited[1])
        info_dict['repetition'] = int(splited[2].split('.')[0])
        return info_dict


class DH_LSA64_pre(DatasetHandler):

    def get_my_url(self):
        return 'https://drive.google.com/uc?id=1yhfPpI2iJzPXyx4C7MYR6IPZC3YuuYaL'

    def get_my_folder(self):
        return 'lsa64_hand_videos'

    def get_my_file_ext(self):
        return 'avi'
    def get_pos_url(self):
        return 'https://drive.google.com/uc?id=1JYTZW8432UAqIm2tXrLNOASzZwdj3EDH'    

    def video_info(self, filename):
        splited = filename.split('_')
        info_dict = {}
        info_dict['class']= int(splited[0])
        info_dict['consultant']= int(splited[1])
        info_dict['repetition'] = int(splited[2])
        info_dict['hand'] = splited[3].split('.')[0]
        return info_dict

class DH_ASLLVD_pre(DatasetHandler):

    def set_dai_path(self, mypath):
        from pandas import read_excel as rex        
        dai_path = osp.join(mypath, 'dai.xlsx')
        if not osp.exists(dai_path):
            self.get_dai(dai_path)
        self.dframe=rex(dai_path)

    def get_my_url(self):
        return 'http://csr.bu.edu/ftp/asl/asllvd/asl-data2/quicktime/'

    def get_my_file_ext(self):
        return 'mov'

    def get_pos_url(self):
        raise NotImplementedError

    def get_first(self):        
        session, scene = self.mov_list()[2]
        return self.get_my_url() + session + '/scene' + str(scene)+'-camera1.mov'

    def get_dai(self, dpath):
        import requests
        import openpyxl as pyxl
        from tempfile import NamedTemporaryFile as nt
        resp = requests.get(
            'http://www.bu.edu/asllrp/dai-asllvd-BU_glossing_with_variations_HS_information-extended-urls-RU.xlsx')
        t=nt()    
        with open(t.name, 'wb') as f:
            f.write(resp.content)
            f.close()
        new_name=t.name + '.xlsx'
        rename(t.name,new_name)
        wb = pyxl.load_workbook(new_name)
        ws = wb.active        
        for link in filter(lambda cell: cell.value.startswith('=HYPER') ,ws['L']):
            link.value = link.value.split('"')[1].split('/')[-1]
        ws['L'][0].value='Filename'
        ws['D'][0].value='Main_Gloss'
        wb.save(dpath)
        


    def mov_list(self):        
        return self.dframe[['Session', 'Scene']].drop_duplicates().sort_values(['Session', 'Scene']).get_values()

    def get_urls(self):
        urls = []
        for session, scene in (self.mov_list()[2:]):
            urls.append(self.get_my_url() + session +
                        '/scene' + str(scene)+'-camera1.mov')
        return urls  
    
    def video_info(self, filename):        
        return self.dframe.loc[:, 'Consultant':'Passive Arm'][(self.dframe.Filename == filename)].to_dict('records')[0]
        
    def consultant_of(self, d):
        return d['Consultant'].lower()

    def word_class(self, d):
        return d['Main_Gloss'].lower()

    def rept(self, d):
        return None

    def extract_signs_from(self, scene_path):
        from skvideo.io import vread            
        session, filename = scene_path.split('/')[-2:]
        num=int(filename.split('-')[0][-1])        
        sign_time=self.dframe[(self.dframe['Session'] == session)&(self.dframe['Scene']== num )][['Filename','Start','End']].get_values()
        video = vread(scene_path)
        d={}
        for _l,name, start, end in sign_time:            
            d[name]=video[slice(start,end)]        
        return d
